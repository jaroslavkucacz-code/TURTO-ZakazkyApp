"""Excel Ceník parser for Nevoga customer and transport variants."""
from __future__ import annotations
import json,re
from pathlib import Path
from .common import _norm,_number
from .model import _base_item

def _header_map(values: list) -> dict:
    return {_norm(value): index for index, value in enumerate(values) if str(value or "").strip()}


def _cell(row, mapping, *names):
    for name in names:
        index = mapping.get(_norm(name))
        if index is not None and index < len(row):
            return row[index]
    return None


def parse_excel_price_list(path: Path, progress=None) -> dict:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Pro import Excel ceníků chybí knihovna openpyxl.") from exc
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if progress:progress(0,max(1,len(rows)),f"Načteno {len(rows)} řádků z Excelu")
    if not rows:
        return {"supplier": "", "title": path.stem, "items": [], "parse_status": "Prázdný sešit", "source_type": "XLSX"}
    header_index = 0
    for idx, row in enumerate(rows[:12]):
        normed = {_norm(value) for value in row if value not in (None, "")}
        if {"article number", "net price per unit"}.issubset(normed) or {"item no", "price czk"}.issubset(normed):
            header_index = idx
            break
    header = list(rows[header_index])
    mapping = _header_map(header)
    items = []
    kind = "generic"
    if "article number" in mapping and "net price per unit" in mapping:
        kind = "customer_specific"
        for row_no, row in enumerate(rows[header_index + 1:], header_index + 2):
            if progress and row_no % 100 == 0:progress(row_no,len(rows),f"Zpracovávám řádek {row_no} z {len(rows)}")
            code = str(_cell(row, mapping, "Article number") or "").strip()
            name = str(_cell(row, mapping, "Designation 1") or "").strip()
            if not code and not name:
                continue
            description = " | ".join(str(x).strip() for x in (
                _cell(row, mapping, "Designation 2"), _cell(row, mapping, "Dimensionstext"),
                _cell(row, mapping, "Longtext")) if str(x or "").strip())
            discount_raw = _number(_cell(row, mapping, "Discount rate"))
            discount_pct = discount_raw * 100 if 0 <= discount_raw <= 1 else discount_raw
            item = _base_item(
                row_no=row_no, product_code=code, name=name, description=description,
                unit=str(_cell(row, mapping, "Sales quantity unit") or ""),
                source_price=_number(_cell(row, mapping, "gross list price")),
                normalized_unit_price=_number(_cell(row, mapping, "Net price per unit")),
                discount_pct=discount_pct, minimum_qty=_number(_cell(row, mapping, "From quantity")),
                package_qty=_number(_cell(row, mapping, "Conversion factor VPVK")),
                package_unit=str(_cell(row, mapping, "Packaging unit") or ""),
                gtin=str(_cell(row, mapping, "GTIN") or ""),
                customs_code=str(_cell(row, mapping, "Customs tariff") or ""),
                weight_unit=_number(_cell(row, mapping, "Weight baseunit")),
                weight_package=_number(_cell(row, mapping, "Weight sale quantity unit")),
                source_row_json=json.dumps({str(header[i] or i): row[i] for i in range(min(len(header), len(row)))}, ensure_ascii=False, default=str),
            )
            items.append(item)
    elif "item no" in mapping and "price czk" in mapping:
        kind = "nevoga"
        has_new_price = "nova nakupni cena" in mapping
        for row_no, row in enumerate(rows[header_index + 1:], header_index + 2):
            if progress and row_no % 100 == 0:progress(row_no,len(rows),f"Zpracovávám řádek {row_no} z {len(rows)}")
            code = str(_cell(row, mapping, "Item-No.") or "").strip()
            name = str(_cell(row, mapping, "Product description") or "").strip()
            if not code and not name:
                continue
            source_price = _number(_cell(row, mapping, "Price CZK"))
            basis = _number(_cell(row, mapping, "Price CZK per"), 1.0) or 1.0
            adjustment_cols = [i for key, i in mapping.items() if key == _norm("Discount / Addition %")]
            # duplicate headers are collapsed by dict; recover from raw header positions
            adjustment_cols = [i for i, value in enumerate(header) if _norm(value) == _norm("Discount / Addition %")]
            adjustment = sum(_number(row[i]) for i in adjustment_cols if i < len(row))
            normalized = _number(_cell(row, mapping, "Nová nákupní cena")) if has_new_price else 0.0
            if not normalized:
                normalized = source_price * (1 + adjustment / 100.0) / basis
            surcharge = max(adjustment, 0)
            discount = max(-adjustment, 0)
            attrs = []
            for key, label, unit in (
                ("Dimensions / pal.", "Rozměry palety", ""),
                ("Dimensions / PU1", "Rozměry balení 1", ""),
                ("Dimensions / PU2", "Rozměry balení 2", ""),
            ):
                value = str(_cell(row, mapping, key) or "").strip()
                if value and value not in {"-", "—"}:
                    attrs.append({"key": label, "value": value, "unit": unit, "source": "XLSX"})
            item = _base_item(
                row_no=row_no, product_code=code, supplier_item_code=str(_cell(row, mapping, "Your Item-No.") or ""),
                name=name, description=name, unit=str(_cell(row, mapping, "Unit") or ""),
                source_price=source_price, price_basis_qty=basis, normalized_unit_price=normalized,
                discount_pct=discount, surcharge_pct=surcharge,
                minimum_qty=_number(_cell(row, mapping, "Price CZK starting from (units)")),
                pallet_qty=_number(_cell(row, mapping, "Units/pal.")),
                package_qty=_number(_cell(row, mapping, "Units per PU1 (= minimum order quantity)")),
                package_unit=str(_cell(row, mapping, "Packing Unit 1 (PU1 = minimum order unit)") or ""),
                weight_unit=_number(_cell(row, mapping, "Weight in kgs per Unit")),
                weight_pallet=_number(_cell(row, mapping, "Weight in kgs per pal.")),
                weight_package=_number(_cell(row, mapping, "Weight in kgs per PU1")),
                customs_code=str(_cell(row, mapping, "Customs tariff number") or ""),
                dimensions=str(_cell(row, mapping, "Dimensions / pal.") or ""),
                condition_text=(f"Příplatek {adjustment:g} %" if adjustment > 0 else
                                f"Sleva {abs(adjustment):g} %" if adjustment < 0 else ""),
                source_row_json=json.dumps({str(header[i] or i): row[i] for i in range(min(len(header), len(row)))}, ensure_ascii=False, default=str),
                attributes=attrs,
            )
            items.append(item)
    else:
        # Preserve the document; unknown layouts remain visible in evidence.
        items = []
    title = path.stem
    valid_from = ""
    date_match = re.search(r"(\d{1,2})[.\-_](\d{1,2})[.\-_](20\d{2})", path.stem)
    if date_match:
        valid_from = f"{date_match.group(3)}-{int(date_match.group(2)):02d}-{int(date_match.group(1)):02d}"
    product_group = "BETONY" if "BETON" in path.stem.upper() else ""
    if product_group:
        branch = "Polsko – ucelený kamion"
    elif kind == "nevoga":
        branch = "Česká republika – neucelená dodávka"
    elif kind == "customer_specific":
        branch = "Česká republika – zákaznický ceník"
    else:
        branch = ""
    return {
        "supplier": "Nevoga" if kind in {"nevoga", "customer_specific"} else "",
        "title": title, "valid_from": valid_from, "valid_to": "",
        "product_group": product_group, "branch": branch, "currency": "CZK",
        "items": items, "terms_text": "", "raw_text": "", "ocr_text": "", "ocr_layout_json": "", "ocr_engine": "",
        "parse_status": "Rozpoznáno automaticky" if items else "Bez rozpoznaných položek",
        "source_type": "XLSX", "suggested_update_mode": "partial" if product_group else "replace_all",
    }
